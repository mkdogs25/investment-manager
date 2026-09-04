"""Builds the .xlsx report with openpyxl.

Five sheets: Portfolio Summary, Fund Analysis, Fund Details, Methodology and
Data Sources.  Anything the application could not retrieve is written as the
literal string "Data unavailable" so that it is never confused with zero.
"""

from __future__ import annotations

import io
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ..config import settings
from ..models.schemas import AnalysisResponse, FundAnalysis
from .analysis_service import DISCLAIMER
from .scoring import MIN_COMPONENTS, methodology_rows

UNAVAILABLE = "Data unavailable"

# --- palette ----------------------------------------------------------
INK = "1F2937"
BRAND = "1E3A5F"
ACCENT = "0F766E"
MUTED = "6B7280"
BAND = "F1F5F9"
HEADER_FILL = PatternFill("solid", fgColor=BRAND)
SUBHEAD_FILL = PatternFill("solid", fgColor="E2E8F0")
LABEL_FILL = PatternFill("solid", fgColor="F8FAFC")

TITLE_FONT = Font(name="Calibri", size=18, bold=True, color=BRAND)
SUBTITLE_FONT = Font(name="Calibri", size=10, italic=True, color=MUTED)
SECTION_FONT = Font(name="Calibri", size=12, bold=True, color=ACCENT)
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
LABEL_FONT = Font(name="Calibri", size=11, bold=True, color=INK)
BODY_FONT = Font(name="Calibri", size=11, color=INK)
NOTE_FONT = Font(name="Calibri", size=9, italic=True, color=MUTED)

THIN = Side(style="thin", color="D1D5DB")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# --- number formats ---------------------------------------------------
# ``#,##,##0.00`` is the Indian lakh/crore grouping mask: 15000000 renders as
# 1,50,00,000.00. Two plain sections are used rather than conditional ones so
# that the minus sign on negative amounts is always displayed.
INR = '"₹"#,##,##0.00'
INR_SIGNED = '"₹"#,##,##0.00;[Red]-"₹"#,##,##0.00'
PCT = '0.00"%";[Red]-0.00"%"'
NUM2 = "0.00"
DATE_FMT = "DD-MM-YYYY"


# ----------------------------------------------------------------------
# helpers
# ----------------------------------------------------------------------
def _write(
    ws: Worksheet,
    row: int,
    col: int,
    value,
    *,
    font: Font | None = None,
    fill: PatternFill | None = None,
    number_format: str | None = None,
    align: str | None = None,
    border: bool = False,
    wrap: bool = False,
):
    cell = ws.cell(row=row, column=col)
    cell.value = UNAVAILABLE if value is None else value
    cell.font = font or BODY_FONT
    if fill:
        cell.fill = fill
    if number_format and value is not None:
        cell.number_format = number_format
    cell.alignment = Alignment(
        horizontal=align or ("left" if value is None or isinstance(value, str) else "right"),
        vertical="center",
        wrap_text=wrap,
    )
    if border:
        cell.border = BOX
    return cell


def _title(ws: Worksheet, row: int, text: str, subtitle: str | None = None) -> int:
    _write(ws, row, 1, text, font=TITLE_FONT, align="left")
    row += 1
    if subtitle:
        _write(ws, row, 1, subtitle, font=SUBTITLE_FONT, align="left")
        row += 1
    return row + 1


def _section(ws: Worksheet, row: int, text: str) -> int:
    _write(ws, row, 1, text, font=SECTION_FONT, align="left")
    return row + 1


def _header_row(ws: Worksheet, row: int, headers: list[str]) -> int:
    for idx, header in enumerate(headers, start=1):
        _write(
            ws,
            row,
            idx,
            header,
            font=HEADER_FONT,
            fill=HEADER_FILL,
            align="center",
            border=True,
            wrap=True,
        )
    ws.row_dimensions[row].height = 32
    return row + 1


def _widths(ws: Worksheet, widths: list[int]) -> None:
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _disclaimer(ws: Worksheet, row: int, span: int = 8) -> int:
    row += 1
    _write(ws, row, 1, "Disclaimer", font=SECTION_FONT, align="left")
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 2, end_column=max(span, 4))
    cell = _write(ws, row, 1, DISCLAIMER, font=NOTE_FONT, align="left", wrap=True)
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    return row + 4


def _fund_label(fund: FundAnalysis) -> str:
    if fund.scheme and fund.scheme.scheme_name:
        return fund.scheme.scheme_name
    return fund.input.fund_name


# ----------------------------------------------------------------------
# Sheet 1 — Portfolio Summary
# ----------------------------------------------------------------------
def _sheet_summary(wb: Workbook, report: AnalysisResponse) -> None:
    ws = wb.active
    ws.title = "Portfolio Summary"
    ws.sheet_view.showGridLines = False
    _widths(ws, [38, 20, 20, 20, 16, 16, 18])

    row = _title(
        ws,
        1,
        "Mutual Fund Portfolio Analyzer",
        "Portfolio summary — figures in INR unless stated otherwise",
    )

    row = _section(ws, row, "Portfolio Summary")
    totals = report.totals
    summary_rows: list[tuple[str, object, str | None]] = [
        ("Investor Age", report.investor_age, "0"),
        ("Risk Profile", report.risk_profile.value, None),
        ("Number of Funds", totals.number_of_funds, "0"),
        ("Total Invested", totals.total_invested, INR),
        ("Current Portfolio Value", totals.total_current, INR),
        ("Total Gain/Loss", totals.total_gain_loss, INR_SIGNED),
        ("Total Return", totals.total_return_pct, PCT),
        ("Report Generated", report.generated_at, None),
        ("Largest Holding", totals.largest_holding, None),
        ("Largest Holding Value", totals.largest_holding_value, INR),
        ("Smallest Holding", totals.smallest_holding, None),
        ("Smallest Holding Value", totals.smallest_holding_value, INR),
    ]
    for label, value, fmt in summary_rows:
        _write(ws, row, 1, label, font=LABEL_FONT, fill=LABEL_FILL, align="left", border=True)
        _write(ws, row, 2, value, number_format=fmt, border=True)
        row += 1

    gain_loss_row = row - 12 + 5
    ws.conditional_formatting.add(
        f"B{gain_loss_row}:B{gain_loss_row + 1}",
        CellIsRule(operator="lessThan", formula=["0"], font=Font(color="B91C1C", bold=True)),
    )
    ws.conditional_formatting.add(
        f"B{gain_loss_row}:B{gain_loss_row + 1}",
        CellIsRule(operator="greaterThan", formula=["0"], font=Font(color="15803D", bold=True)),
    )

    row += 1
    row = _section(ws, row, "Portfolio Allocation")
    headers = ["Fund", "Invested", "Current Value", "Gain/Loss", "Return %", "Allocation %"]
    header_row = row
    row = _header_row(ws, row, headers)

    first_data_row = row
    for fund in report.funds:
        _write(ws, row, 1, _fund_label(fund), align="left", border=True)
        _write(ws, row, 2, fund.input.amount_invested, number_format=INR, border=True)
        _write(ws, row, 3, fund.input.current_amount, number_format=INR, border=True)
        _write(ws, row, 4, fund.gain_loss, number_format=INR_SIGNED, border=True)
        _write(ws, row, 5, fund.absolute_return_pct, number_format=PCT, border=True)
        _write(ws, row, 6, fund.allocation_pct, number_format=PCT, border=True)
        row += 1
    last_data_row = row - 1

    _write(ws, row, 1, "TOTAL", font=LABEL_FONT, fill=SUBHEAD_FILL, align="left", border=True)
    _write(ws, row, 2, totals.total_invested, font=LABEL_FONT, fill=SUBHEAD_FILL,
           number_format=INR, border=True)
    _write(ws, row, 3, totals.total_current, font=LABEL_FONT, fill=SUBHEAD_FILL,
           number_format=INR, border=True)
    _write(ws, row, 4, totals.total_gain_loss, font=LABEL_FONT, fill=SUBHEAD_FILL,
           number_format=INR_SIGNED, border=True)
    _write(ws, row, 5, totals.total_return_pct, font=LABEL_FONT, fill=SUBHEAD_FILL,
           number_format=PCT, border=True)
    _write(ws, row, 6, 100.0 if totals.total_current > 0 else None, font=LABEL_FONT,
           fill=SUBHEAD_FILL, number_format=PCT, border=True)
    total_row = row
    row += 2

    if last_data_row >= first_data_row:
        ws.freeze_panes = f"A{first_data_row}"
        ws.conditional_formatting.add(
            f"D{first_data_row}:E{last_data_row}",
            CellIsRule(operator="lessThan", formula=["0"], font=Font(color="B91C1C")),
        )
        ws.conditional_formatting.add(
            f"D{first_data_row}:E{last_data_row}",
            CellIsRule(operator="greaterThan", formula=["0"], font=Font(color="15803D")),
        )
        ws.conditional_formatting.add(
            f"F{first_data_row}:F{last_data_row}",
            ColorScaleRule(
                start_type="min", start_color="FFFFFF",
                end_type="max", end_color="93C5FD",
            ),
        )

        chart = PieChart()
        chart.title = "Portfolio Allocation by Current Value"
        chart.height = 9
        chart.width = 16
        labels = Reference(ws, min_col=1, min_row=first_data_row, max_row=last_data_row)
        values = Reference(ws, min_col=3, min_row=header_row, max_row=last_data_row)
        chart.add_data(values, titles_from_data=True)
        chart.set_categories(labels)
        ws.add_chart(chart, f"H{header_row}")

    row = _section(ws, row, "Risk Profile Context")
    _write(
        ws, row, 1,
        f"Investor profile: {report.risk_profile.value}. "
        "This is recorded for context only — the report makes no buy, sell or switch "
        "recommendation. The figures below describe the portfolio as it stands.",
        font=NOTE_FONT, align="left", wrap=True,
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 1, end_column=6)
    row += 3

    largest_alloc = max(
        (f.allocation_pct for f in report.funds if f.allocation_pct is not None),
        default=None,
    )
    top_three = sum(
        sorted((f.allocation_pct or 0.0) for f in report.funds)[-3:]
    ) if report.funds else None
    context_rows: list[tuple[str, object, str | None]] = [
        ("Number of holdings", totals.number_of_funds, "0"),
        ("Largest single-fund allocation", largest_alloc, PCT),
        ("Top 3 holdings as share of portfolio", top_three, PCT),
        (
            "Funds with complete scheme data",
            sum(1 for f in report.funds if f.status.value == "retrieved"),
            "0",
        ),
    ]
    for label, value, fmt in context_rows:
        _write(ws, row, 1, label, font=LABEL_FONT, fill=LABEL_FILL, align="left", border=True)
        _write(ws, row, 2, value, number_format=fmt, border=True)
        row += 1

    row += 1
    row = _section(ws, row, "Category Distribution (by current value)")
    row = _header_row(ws, row, ["Scheme Category", "Share of Portfolio"])
    if report.category_distribution:
        for category, share in report.category_distribution.items():
            _write(ws, row, 1, category, align="left", border=True)
            _write(ws, row, 2, share, number_format=PCT, border=True)
            row += 1
    else:
        _write(ws, row, 1, UNAVAILABLE, align="left", border=True)
        _write(ws, row, 2, None, border=True)
        row += 1

    _disclaimer(ws, row, span=6)
    _ = total_row


# ----------------------------------------------------------------------
# Sheet 2 — Fund Analysis
# ----------------------------------------------------------------------
FUND_ANALYSIS_HEADERS = [
    "Fund", "AMC", "Category", "Scheme Type", "Investment Date", "Holding Period",
    "Amount Invested", "Current Amount", "Gain/Loss", "Return %", "CAGR",
    "1Y Return", "3Y Return", "5Y Return", "Volatility", "Sharpe",
    "Max Drawdown", "Fund Score", "Allocation %", "Data Status",
]


def _sheet_fund_analysis(wb: Workbook, report: AnalysisResponse) -> None:
    ws = wb.create_sheet("Fund Analysis")
    ws.sheet_view.showGridLines = False
    _widths(ws, [42, 24, 24, 22, 15, 22, 16, 16, 16, 12, 12, 12, 12, 12, 12, 10, 14, 12, 13, 18])

    row = _title(
        ws, 1, "Fund Analysis",
        "Trailing returns, risk metrics and the Fund Analysis Score are calculated by "
        "this application from published NAV history — see the Methodology sheet.",
    )
    header_row = row
    row = _header_row(ws, row, FUND_ANALYSIS_HEADERS)
    first_data_row = row

    for fund in report.funds:
        scheme = fund.scheme
        _write(ws, row, 1, _fund_label(fund), align="left", border=True)
        _write(ws, row, 2, scheme.fund_house if scheme else None, align="left", border=True)
        _write(ws, row, 3, scheme.scheme_category if scheme else None, align="left", border=True)
        _write(ws, row, 4, scheme.scheme_type if scheme else None, align="left", border=True)
        _write(ws, row, 5, fund.input.investment_date, number_format=DATE_FMT,
               align="center", border=True)
        _write(
            ws, row, 6,
            fund.holding_period_label if fund.holding_period_days is not None else None,
            align="left", border=True,
        )
        _write(ws, row, 7, fund.input.amount_invested, number_format=INR, border=True)
        _write(ws, row, 8, fund.input.current_amount, number_format=INR, border=True)
        _write(ws, row, 9, fund.gain_loss, number_format=INR_SIGNED, border=True)
        _write(ws, row, 10, fund.absolute_return_pct, number_format=PCT, border=True)
        cagr_cell = _write(ws, row, 11, fund.cagr_pct, number_format=PCT, border=True)
        if fund.cagr_pct is None and fund.cagr_note:
            cagr_cell.value = fund.cagr_note
            cagr_cell.font = NOTE_FONT
            cagr_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        _write(ws, row, 12, fund.return_1y.value, number_format=PCT, border=True)
        _write(ws, row, 13, fund.return_3y.value, number_format=PCT, border=True)
        _write(ws, row, 14, fund.return_5y.value, number_format=PCT, border=True)
        _write(ws, row, 15, fund.volatility.value, number_format=PCT, border=True)
        _write(ws, row, 16, fund.sharpe_ratio.value, number_format=NUM2, border=True)
        _write(ws, row, 17, fund.max_drawdown.value, number_format=PCT, border=True)
        score_cell = _write(ws, row, 18, fund.fund_score, number_format=NUM2, border=True)
        if fund.fund_score is None:
            score_cell.value = "Score unavailable — insufficient data"
            score_cell.font = NOTE_FONT
            score_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        _write(ws, row, 19, fund.allocation_pct, number_format=PCT, border=True)
        _write(ws, row, 20, fund.status_label, align="left", border=True)
        row += 1

    last_data_row = row - 1
    ws.freeze_panes = f"A{first_data_row}"
    if last_data_row >= first_data_row:
        ws.auto_filter.ref = f"A{header_row}:T{last_data_row}"
        for column in ("I", "J", "K", "L", "M", "N", "Q"):
            ws.conditional_formatting.add(
                f"{column}{first_data_row}:{column}{last_data_row}",
                CellIsRule(operator="lessThan", formula=["0"], font=Font(color="B91C1C")),
            )
            ws.conditional_formatting.add(
                f"{column}{first_data_row}:{column}{last_data_row}",
                CellIsRule(operator="greaterThan", formula=["0"], font=Font(color="15803D")),
            )
        ws.conditional_formatting.add(
            f"R{first_data_row}:R{last_data_row}",
            ColorScaleRule(
                start_type="num", start_value=0, start_color="FCA5A5",
                mid_type="num", mid_value=50, mid_color="FEF3C7",
                end_type="num", end_value=100, end_color="86EFAC",
            ),
        )

    row += 1
    _write(
        ws, row, 1,
        "Fund Score is this application's own transparent score (0-100). It is not a "
        "rating from Value Research or any other organisation.",
        font=NOTE_FONT, align="left",
    )
    row += 1

    messages = [(f, m) for f in report.funds for m in f.messages]
    if messages:
        row += 1
        row = _section(ws, row, "Per-fund notes")
        for fund, message in messages:
            _write(ws, row, 1, _fund_label(fund), align="left", border=True)
            _write(ws, row, 2, message, align="left", border=True, wrap=True)
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
            row += 1

    _disclaimer(ws, row, span=8)


# ----------------------------------------------------------------------
# Sheet 3 — Fund Details
# ----------------------------------------------------------------------
def _sheet_fund_details(wb: Workbook, report: AnalysisResponse) -> None:
    ws = wb.create_sheet("Fund Details")
    ws.sheet_view.showGridLines = False
    _widths(ws, [30, 60])

    row = _title(
        ws, 1, "Fund Details",
        "Scheme identification and metadata exactly as published by the data provider.",
    )

    for fund in report.funds:
        row = _section(ws, row, _fund_label(fund))
        scheme = fund.scheme
        fields: list[tuple[str, object, str | None]] = [
            ("Scheme Name", scheme.scheme_name if scheme else None, None),
            ("AMC", scheme.fund_house if scheme else None, None),
            ("Scheme Code", scheme.scheme_code if scheme else None, None),
            ("ISIN (Growth)", scheme.isin_growth if scheme else None, None),
            ("ISIN (IDCW Reinvestment)", scheme.isin_div_reinvestment if scheme else None, None),
            ("Category", scheme.scheme_category if scheme else None, None),
            ("Scheme Type", scheme.scheme_type if scheme else None, None),
            ("Direct/Regular", scheme.plan if scheme else None, None),
            ("Growth/IDCW", scheme.option if scheme else None, None),
            ("Latest NAV", scheme.latest_nav if scheme else None, NUM2),
            ("NAV Date", scheme.latest_nav_date if scheme else None, DATE_FMT),
            (
                "Earliest Published NAV Date",
                scheme.inception_date if scheme else None,
                DATE_FMT,
            ),
            ("NAV Observations Used", scheme.nav_history_points if scheme else None, "0"),
            ("AUM", fund.aum.value, INR),
            ("Expense Ratio", fund.expense_ratio.value, PCT),
            ("Benchmark", fund.benchmark, None),
            ("Data Status", fund.status_label, None),
        ]
        for label, value, fmt in fields:
            # Only render a field when it has a value, or when its absence is
            # meaningful to the reader (identification and status fields).
            always = label in {
                "Scheme Name", "AMC", "Scheme Code", "Category",
                "Latest NAV", "NAV Date", "Data Status",
            }
            if value is None and not always:
                continue
            _write(ws, row, 1, label, font=LABEL_FONT, fill=LABEL_FILL, align="left", border=True)
            _write(ws, row, 2, value, number_format=fmt, align="left", border=True)
            row += 1

        if fund.messages:
            _write(ws, row, 1, "Notes", font=LABEL_FONT, fill=LABEL_FILL, align="left",
                   border=True)
            _write(ws, row, 2, " ".join(fund.messages), font=NOTE_FONT, align="left",
                   border=True, wrap=True)
            row += 1
        row += 1

    _disclaimer(ws, row, span=2)


# ----------------------------------------------------------------------
# Sheet 4 — Methodology
# ----------------------------------------------------------------------
def _sheet_methodology(wb: Workbook, report: AnalysisResponse) -> None:
    ws = wb.create_sheet("Methodology")
    ws.sheet_view.showGridLines = False
    _widths(ws, [32, 100])

    row = _title(
        ws, 1, "Methodology",
        "Every figure in this workbook is reproducible from the formulae below.",
    )

    rf = settings.risk_free_rate * 100
    entries: list[tuple[str, str]] = [
        (
            "Absolute Gain/Loss",
            "Current Amount − Amount Invested, using the values you entered. Your inputs "
            "are never adjusted.",
        ),
        (
            "Absolute Return %",
            "(Current Amount − Amount Invested) ÷ Amount Invested × 100.",
        ),
        (
            "Holding Period",
            "Exact elapsed calendar time between the investment date and the report "
            "generation date, expressed in years, months and days.",
        ),
        (
            "CAGR",
            "((Current Amount ÷ Amount Invested) ^ (365.25 ÷ holding period in days)) − 1, "
            "× 100. The actual holding period is used — no rounding to whole years. CAGR is "
            "not calculated when the investment date is missing or invalid, when the holding "
            "period is under one day, or when the current value is zero.",
        ),
        (
            "Portfolio Allocation %",
            "Current Fund Value ÷ Total Current Portfolio Value × 100.",
        ),
        (
            "Portfolio Totals",
            "Simple sums of the amounts you entered. Total Return % is total gain/loss ÷ "
            "total invested × 100.",
        ),
        (
            "1Y / 3Y / 5Y Returns",
            "Calculated by this application from the scheme's published NAV history. The 1-year "
            "figure is an absolute return; the 3-year and 5-year figures are annualised "
            "(CAGR). The window start uses the nearest NAV published on or before the target "
            "date, within a 10-day tolerance for market holidays; outside that tolerance the "
            "figure is reported as unavailable. A period is left blank when the NAV history "
            "does not cover it.",
        ),
        (
            "Since-Inception Return",
            "Annualised return from the earliest published NAV to the latest, shown only when "
            "at least one year of history exists. Note that the earliest published NAV date is "
            "not necessarily the scheme's official inception date.",
        ),
        (
            "Volatility",
            "Standard deviation (sample, n−1) of daily NAV returns, annualised by multiplying "
            f"by √{settings.trading_days_per_year}, expressed in percent. Requires at least "
            f"{settings.min_observations_for_risk} daily observations.",
        ),
        (
            "Sharpe Ratio",
            "(annualised return − risk-free rate) ÷ annualised volatility, where the "
            "annualised return is derived from the mean daily NAV return over the full "
            f"available history and the risk-free rate is {rf:.2f}% p.a. (configurable). Both "
            "legs come from the same NAV series.",
        ),
        (
            "Maximum Drawdown",
            "Largest peak-to-trough decline of the NAV series over its full published history, "
            "expressed as a negative percentage.",
        ),
        (
            "Consistency",
            "Share of rolling 1-year windows, sampled monthly, in which the NAV rose. Reported "
            "only when at least 12 such windows exist.",
        ),
        (
            "Fund Analysis Score",
            "A transparent 0-100 score computed by this application from the components in the "
            "table below. Each component is mapped to 0-100 by linear interpolation between "
            "two documented anchor points and clipped to that range; the score is the weighted "
            "mean of the components that could be computed, with weights renormalised over "
            f"those components. Fewer than {MIN_COMPONENTS} available components yields "
            '"Score unavailable — insufficient data". This score is NOT a rating from Value '
            "Research or any other organisation, and it is not a recommendation.",
        ),
        (
            "Data Retrieval",
            f"Scheme identification, metadata and NAV history are retrieved from "
            f"{_provider_label(report)} over its public API. No website is scraped and no "
            "access control is bypassed. Retrieval dates and per-data-point outcomes are "
            "recorded on the Data Sources sheet.",
        ),
        (
            "Unavailable Data",
            'Anything that could not be retrieved or computed is shown as "Data unavailable" '
            "and is never substituted with an estimate, a placeholder or zero.",
        ),
    ]

    row = _header_row(ws, row, ["Metric", "How it is calculated"])
    for label, description in entries:
        _write(ws, row, 1, label, font=LABEL_FONT, fill=LABEL_FILL, align="left", border=True)
        cell = _write(ws, row, 2, description, align="left", border=True, wrap=True)
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[row].height = max(30, 13 * (len(description) // 95 + 1))
        row += 1

    row += 1
    row = _section(ws, row, "Fund Analysis Score — components and weights")
    row = _header_row(ws, row, ["Component", "Weight", "Scale"])
    ws.column_dimensions["C"].width = 90
    for component, weight, description in methodology_rows():
        _write(ws, row, 1, component, font=LABEL_FONT, fill=LABEL_FILL, align="left", border=True)
        _write(ws, row, 2, weight, align="center", border=True)
        cell = _write(ws, row, 3, description, align="left", border=True, wrap=True)
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        row += 1

    row += 1
    row = _section(ws, row, "Assumptions")
    assumptions = [
        "One year is treated as 365.25 days so that leap years do not distort annualisation.",
        f"{settings.trading_days_per_year} trading days per year are assumed when annualising "
        "daily volatility.",
        f"The risk-free rate used by the Sharpe ratio is {rf:.2f}% per annum.",
        "Amounts invested are treated as a single lump sum on the stated investment date; SIP "
        "instalments and additional purchases are not modelled, so the CAGR shown is the "
        "point-to-point return on the figures you entered, not an XIRR.",
        "NAV values of zero in the source feed are treated as missing rather than as prices.",
        "Scheme-level returns are derived from the scheme's own NAV and are independent of "
        "when you personally invested.",
        "Direct/Regular and Growth/IDCW are read from the AMFI scheme name; where the name does "
        'not state them, the field shows "Data unavailable".',
        "All calculations run at full floating-point precision; rounding happens only when a "
        "value is displayed.",
    ]
    for text in assumptions:
        _write(ws, row, 1, "•", align="center", border=False)
        cell = _write(ws, row, 2, text, align="left", wrap=True)
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        row += 1

    _disclaimer(ws, row, span=3)


def _provider_label(report: AnalysisResponse) -> str:
    for record in report.data_sources:
        if record.source not in {"User input", None}:
            return record.source
    return "the configured mutual fund data provider"


# ----------------------------------------------------------------------
# Sheet 5 — Data Sources
# ----------------------------------------------------------------------
def _sheet_data_sources(wb: Workbook, report: AnalysisResponse) -> None:
    ws = wb.create_sheet("Data Sources")
    ws.sheet_view.showGridLines = False
    _widths(ws, [42, 32, 22, 18, 44])

    row = _title(
        ws, 1, "Data Sources",
        "Provenance and retrieval outcome for every externally obtained data point.",
    )
    header_row = row
    row = _header_row(ws, row, ["Fund", "Data Point", "Source", "Retrieved On", "Status"])
    first_data_row = row

    for record in report.data_sources:
        _write(ws, row, 1, record.fund, align="left", border=True)
        _write(ws, row, 2, record.data_point, align="left", border=True)
        _write(ws, row, 3, record.source, align="left", border=True)
        _write(ws, row, 4, record.retrieved_on, align="center", border=True)
        _write(ws, row, 5, record.status, align="left", border=True)
        row += 1

    last_data_row = row - 1
    ws.freeze_panes = f"A{first_data_row}"
    if last_data_row >= first_data_row:
        ws.auto_filter.ref = f"A{header_row}:E{last_data_row}"
        ws.conditional_formatting.add(
            f"E{first_data_row}:E{last_data_row}",
            FormulaRule(
                formula=[f'ISNUMBER(SEARCH("Failed",$E{first_data_row}))'],
                font=Font(color="B91C1C", bold=True),
                fill=PatternFill("solid", bgColor="FEE2E2"),
                stopIfTrue=True,
            ),
        )
        ws.conditional_formatting.add(
            f"E{first_data_row}:E{last_data_row}",
            FormulaRule(
                formula=[f'ISNUMBER(SEARCH("Unavailable",$E{first_data_row}))'],
                font=Font(color="B45309"),
                fill=PatternFill("solid", bgColor="FEF3C7"),
            ),
        )

    row += 1
    _write(
        ws, row, 1,
        "Retrieval dates are recorded in DD-MM-YYYY format. NAV data published by AMFI-derived "
        "sources may lag the market by one business day.",
        font=NOTE_FONT, align="left",
    )
    row += 1
    _disclaimer(ws, row, span=5)


# ----------------------------------------------------------------------
def build_workbook(report: AnalysisResponse) -> Workbook:
    wb = Workbook()
    wb.properties.title = "Mutual Fund Portfolio Analysis"
    wb.properties.creator = "Mutual Fund Portfolio Analyzer"
    wb.properties.created = datetime.now()

    _sheet_summary(wb, report)
    _sheet_fund_analysis(wb, report)
    _sheet_fund_details(wb, report)
    _sheet_methodology(wb, report)
    _sheet_data_sources(wb, report)
    return wb


def build_workbook_bytes(report: AnalysisResponse) -> bytes:
    buffer = io.BytesIO()
    build_workbook(report).save(buffer)
    return buffer.getvalue()


def suggested_filename(today: date | None = None) -> str:
    today = today or date.today()
    return f"portfolio-analysis-{today.strftime('%Y-%m-%d')}.xlsx"
