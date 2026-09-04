from __future__ import annotations

import io

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.config import settings
from app.main import app
from app.providers import registry
from app.providers.base import ProviderRateLimited, ProviderTimeout, ProviderUnavailable

from .conftest import FakeProvider

PORTFOLIO = {
    "investor_age": 34,
    "risk_profile": "Balanced",
    "funds": [
        {
            "scheme_code": "118989",
            "fund_name": "Example Flexi Cap Fund - Direct Plan - Growth",
            "investment_date": "2023-09-04",
            "amount_invested": 100000,
            "current_amount": 150000,
        },
        {
            "scheme_code": "120503",
            "fund_name": "Example Large Cap Fund - Regular Plan - IDCW",
            "investment_date": "2024-01-15",
            "amount_invested": 50000,
            "current_amount": 40000,
        },
    ],
}


@pytest.fixture
def client(provider: FakeProvider):
    registry.set_provider(settings.provider, provider)
    with TestClient(app) as test_client:
        yield test_client
    registry._instances.pop(settings.provider, None)


def test_health(client: TestClient):
    body = client.get("/api/health").json()
    assert body["status"] == "ok"
    assert body["provider"] == "FakeProvider"
    assert "mfapi" in body["available_providers"]


def test_public_config_exposes_no_secrets(client: TestClient):
    body = client.get("/api/config").json()
    assert set(body) == {"provider", "provider_docs", "risk_free_rate_pct", "scoring_weights"}
    assert sum(body["scoring_weights"].values()) == pytest.approx(1.0)


def test_search_returns_selectable_schemes(client: TestClient):
    results = client.get("/api/funds/search", params={"q": "Example"}).json()
    assert len(results) == 2
    assert {r["scheme_code"] for r in results} == {"118989", "120503"}


def test_search_requires_two_characters(client: TestClient):
    assert client.get("/api/funds/search", params={"q": "a"}).status_code == 422


def test_fund_detail_by_scheme_code(client: TestClient):
    body = client.get("/api/funds/118989").json()
    assert body["scheme_code"] == "118989"
    assert body["plan"] == "Direct"
    assert body["option"] == "Growth"
    assert body["latest_nav"] > 0
    assert body["expense_ratio"] is None  # unavailable, not zero


def test_unknown_scheme_returns_404(client: TestClient):
    assert client.get("/api/funds/000000").status_code == 404


@pytest.mark.parametrize(
    ("error", "expected"),
    [
        (ProviderTimeout("t"), 504),
        (ProviderRateLimited("r"), 429),
        (ProviderUnavailable("u"), 503),
    ],
)
def test_provider_failures_map_to_http_status(
    client: TestClient, provider: FakeProvider, error, expected
):
    provider.fail_with = error
    response = client.get("/api/funds/search", params={"q": "Example"})
    assert response.status_code == expected
    assert response.json()["detail"]


def test_analyze_returns_totals_and_disclaimer(client: TestClient):
    body = client.post("/api/portfolio/analyze", json=PORTFOLIO).json()
    assert body["totals"]["total_current"] == 190000
    assert body["funds"][0]["status_label"] == "✓ Data retrieved"
    assert "informational and educational purposes only" in body["disclaimer"]
    assert body["risk_profile"] == "Balanced"


def test_analyze_rejects_duplicate_schemes(client: TestClient):
    payload = {**PORTFOLIO, "funds": [PORTFOLIO["funds"][0], PORTFOLIO["funds"][0]]}
    response = client.post("/api/portfolio/analyze", json=payload)
    assert response.status_code == 422


def test_analyze_rejects_future_investment_date(client: TestClient):
    payload = {
        **PORTFOLIO,
        "funds": [{**PORTFOLIO["funds"][0], "investment_date": "2099-01-01"}],
    }
    assert client.post("/api/portfolio/analyze", json=payload).status_code == 422


def test_analyze_rejects_empty_portfolio(client: TestClient):
    assert client.post("/api/portfolio/analyze", json={**PORTFOLIO, "funds": []}).status_code == 422


def test_analyze_survives_a_total_provider_outage(client: TestClient, provider: FakeProvider):
    provider.fail_with = ProviderUnavailable("down")
    body = client.post("/api/portfolio/analyze", json=PORTFOLIO).json()
    assert body["totals"]["total_current"] == 190000
    assert all(f["status"] == "unavailable" for f in body["funds"])


def test_report_downloads_a_valid_workbook(client: TestClient):
    response = client.post("/api/portfolio/report", json=PORTFOLIO)
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment" in response.headers["content-disposition"]
    assert ".xlsx" in response.headers["content-disposition"]
    wb = load_workbook(io.BytesIO(response.content))
    assert wb.sheetnames == [
        "Portfolio Summary", "Fund Analysis", "Fund Details", "Methodology", "Data Sources",
    ]


def test_report_values_match_the_analysis_endpoint(client: TestClient):
    analysis = client.post("/api/portfolio/analyze", json=PORTFOLIO).json()
    workbook = load_workbook(
        io.BytesIO(client.post("/api/portfolio/report", json=PORTFOLIO).content)
    )
    ws = workbook["Fund Analysis"]
    header_row = next(
        r[0].row for r in ws.iter_rows(min_col=1, max_col=1) if r[0].value == "Fund"
    )
    for offset, fund in enumerate(analysis["funds"], start=1):
        row = header_row + offset
        assert ws.cell(row=row, column=8).value == pytest.approx(fund["input"]["current_amount"])
        assert ws.cell(row=row, column=10).value == pytest.approx(fund["absolute_return_pct"])
        assert ws.cell(row=row, column=11).value == pytest.approx(fund["cagr_pct"])
        assert ws.cell(row=row, column=19).value == pytest.approx(fund["allocation_pct"])
