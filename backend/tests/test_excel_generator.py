from __future__ import annotations

import io
from datetime import date

import pytest
from openpyxl import load_workbook

from app.models.schemas import FundInput, PortfolioRequest, RiskProfile
from app.providers.base import SchemeData, SchemeMeta
from app.services.analysis_service import DISCLAIMER, AnalysisService
from app.services.excel_generator import (
    FUND_ANALYSIS_HEADERS,
    UNAVAILABLE,
    build_workbook,
    build_workbook_bytes,
    suggested_filename,
)
from app.services.fund_service import FundService

from .conftest import REPORT_DATE, FakeProvider, make_scheme

EXPECTED_SHEETS = [
    "Portfolio Summary",
    "Fund Analysis",
    "Fund Details",
    "Methodology",
    "Data Sources",
]


def _request(funds=None) -> PortfolioRequest:
    return PortfolioRequest(
        investor_age=34,
        risk_profile=RiskProfile.balanced,
        funds=funds
        or [
            FundInput(
                scheme_code="118989",
                fund_name="Example Flexi Cap Fund - Direct Plan - Growth",
                investment_date=date(2023, 9, 4),
                amount_invested=100_000,
                current_amount=150_000,
            ),
            FundInput(
                scheme_code="120503",
                fund_name="Example Large Cap Fund - Regular Plan - IDCW",
                investment_date=date(2024, 1, 15),
                amount_invested=50_000,
                current_amount=40_000,
            ),
        ],
    )


@pytest.fixture
async def report(provider: FakeProvider):
    return await AnalysisService(FundService(provider)).analyse(
        _request(), report_date=REPORT_DATE
    )


@pytest.fixture
def workbook(report):
    return load_workbook(io.BytesIO(build_workbook_bytes(report)))


def _cells(ws) -> list[str]:
    return [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]


def test_workbook_is_generated_and_reopens(report):
    payload = build_workbook_bytes(report)
    assert payload[:2] == b"PK"  # a real zip-backed xlsx
    assert len(payload) > 5_000
    load_workbook(io.BytesIO(payload))


def test_all_expected_sheets_exist(workbook):
    assert workbook.sheetnames == EXPECTED_SHEETS


def test_summary_values_match_the_analysis(workbook, report):
    ws = workbook["Portfolio Summary"]
    values = {}
    for row in ws.iter_rows(min_col=1, max_col=2):
        if row[0].value and row[1].value is not None:
            values[str(row[0].value)] = row[1].value

    assert values["Investor Age"] == report.investor_age
    assert values["Risk Profile"] == "Balanced"
    assert values["Number of Funds"] == 2
    assert values["Total Invested"] == pytest.approx(report.totals.total_invested)
    assert values["Current Portfolio Value"] == pytest.approx(report.totals.total_current)
    assert values["Total Gain/Loss"] == pytest.approx(report.totals.total_gain_loss)
    assert values["Total Return"] == pytest.approx(report.totals.total_return_pct)


def test_allocation_block_matches_and_sums_to_100(workbook, report):
    ws = workbook["Portfolio Summary"]
    header_row = next(
        r[0].row for r in ws.iter_rows(min_col=1, max_col=1) if r[0].value == "Fund"
    )
    allocations = []
    row = header_row + 1
    while ws.cell(row=row, column=1).value not in (None, "TOTAL"):
        allocations.append(ws.cell(row=row, column=6).value)
        row += 1
    assert len(allocations) == len(report.funds)
    assert sum(allocations) == pytest.approx(100.0)


def test_summary_has_an_allocation_chart(workbook):
    assert len(workbook["Portfolio Summary"]._charts) == 1


def test_fund_analysis_headers_and_row_count(workbook, report):
    ws = workbook["Fund Analysis"]
    header_row = next(
        r[0].row for r in ws.iter_rows(min_col=1, max_col=1) if r[0].value == "Fund"
    )
    headers = [ws.cell(row=header_row, column=i + 1).value for i in range(len(FUND_ANALYSIS_HEADERS))]
    assert headers == FUND_ANALYSIS_HEADERS
    for offset, fund in enumerate(report.funds, start=1):
        row = header_row + offset
        assert ws.cell(row=row, column=1).value == fund.scheme.scheme_name
        assert ws.cell(row=row, column=7).value == pytest.approx(fund.input.amount_invested)
        assert ws.cell(row=row, column=8).value == pytest.approx(fund.input.current_amount)
        assert ws.cell(row=row, column=9).value == pytest.approx(fund.gain_loss)
        assert ws.cell(row=row, column=10).value == pytest.approx(fund.absolute_return_pct)
        assert ws.cell(row=row, column=11).value == pytest.approx(fund.cagr_pct)


def test_fund_analysis_formatting_is_applied(workbook):
    ws = workbook["Fund Analysis"]
    header_row = next(
        r[0].row for r in ws.iter_rows(min_col=1, max_col=1) if r[0].value == "Fund"
    )
    first = header_row + 1
    assert ws.freeze_panes == f"A{first}"
    assert ws.auto_filter.ref is not None
    assert ws.auto_filter.ref.startswith(f"A{header_row}")
    assert ws.conditional_formatting
    assert ws.column_dimensions["A"].width > 20
    assert "₹" in ws.cell(row=first, column=7).number_format
    assert "%" in ws.cell(row=first, column=10).number_format
    assert ws.cell(row=first, column=5).number_format == "DD-MM-YYYY"
    assert ws.cell(row=header_row, column=1).font.bold


def test_fund_details_lists_every_scheme(workbook, report):
    text = " ".join(_cells(workbook["Fund Details"]))
    for fund in report.funds:
        assert fund.scheme.scheme_name in text
        assert fund.scheme.scheme_code in text
    assert "ISIN (Growth)" in text
    assert "Direct" in text and "Regular" in text


def test_methodology_documents_every_metric(workbook):
    text = " ".join(_cells(workbook["Methodology"]))
    for topic in (
        "CAGR", "Absolute Return", "Portfolio Allocation", "Volatility",
        "Sharpe Ratio", "Maximum Drawdown", "Fund Analysis Score",
        "Data Retrieval", "Assumptions",
    ):
        assert topic in text
    assert "365.25" in text
    assert "Value Research" in text  # the explicit non-affiliation statement
    for component in ("Long Term Return", "Consistency", "Sharpe", "Expense Ratio"):
        assert component in text


def test_data_sources_records_provenance(workbook, report):
    ws = workbook["Data Sources"]
    header_row = next(
        r[0].row for r in ws.iter_rows(min_col=1, max_col=1) if r[0].value == "Fund"
    )
    rows = [
        [ws.cell(row=r, column=c).value for c in range(1, 6)]
        for r in range(header_row + 1, header_row + 1 + len(report.data_sources))
    ]
    assert len(rows) == len(report.data_sources)
    assert all(row[3] == "04-09-2026" for row in rows)
    assert any(row[1] == "Historical NAV" and row[2] == "FakeProvider" for row in rows)
    assert ws.auto_filter.ref is not None


def test_disclaimer_appears_on_every_sheet(workbook):
    for name in EXPECTED_SHEETS:
        text = " ".join(_cells(workbook[name]))
        assert "informational and educational purposes only" in text
    assert DISCLAIMER.startswith("This tool is for informational")


async def test_missing_data_is_labelled_not_zeroed(provider: FakeProvider):
    thin = SchemeData(
        meta=SchemeMeta(scheme_code="333333", scheme_name="Thin Scheme - Growth"),
        nav_history=make_scheme("333333", "Thin", history_years=0.15).nav_history,
    )
    report = await AnalysisService(FundService(FakeProvider({"333333": thin}))).analyse(
        _request(
            funds=[
                FundInput(
                    scheme_code="333333",
                    fund_name="Thin Scheme - Growth",
                    investment_date=None,
                    amount_invested=1_000,
                    current_amount=1_100,
                )
            ]
        ),
        report_date=REPORT_DATE,
    )
    wb = load_workbook(io.BytesIO(build_workbook_bytes(report)))
    ws = wb["Fund Analysis"]
    header_row = next(
        r[0].row for r in ws.iter_rows(min_col=1, max_col=1) if r[0].value == "Fund"
    )
    row = header_row + 1
    assert ws.cell(row=row, column=2).value == UNAVAILABLE   # AMC
    assert ws.cell(row=row, column=13).value == UNAVAILABLE  # 3Y return
    assert ws.cell(row=row, column=11).value == "CAGR unavailable — investment date required"
    assert ws.cell(row=row, column=18).value == "Score unavailable — insufficient data"
    # The user's own figures are still real numbers, not the unavailable marker.
    assert ws.cell(row=row, column=7).value == 1_000
    assert ws.cell(row=row, column=9).value == 100


async def test_workbook_builds_when_every_fund_fails(outage_provider: FakeProvider):
    report = await AnalysisService(FundService(outage_provider)).analyse(
        _request(), report_date=REPORT_DATE
    )
    wb = load_workbook(io.BytesIO(build_workbook_bytes(report)))
    assert wb.sheetnames == EXPECTED_SHEETS
    text = " ".join(_cells(wb["Data Sources"]))
    assert "Failed" in text


async def test_large_portfolio_workbook(scheme_a: SchemeData):
    provider = FakeProvider({f"{i:06d}": scheme_a for i in range(22)})
    funds = [
        FundInput(
            scheme_code=f"{i:06d}",
            fund_name=f"Fund {i}",
            investment_date=date(2022, 1, 10),
            amount_invested=10_000 + i * 100,
            current_amount=12_000 + i * 100,
        )
        for i in range(22)
    ]
    report = await AnalysisService(FundService(provider)).analyse(
        _request(funds=funds), report_date=REPORT_DATE
    )
    wb = load_workbook(io.BytesIO(build_workbook_bytes(report)))
    ws = wb["Fund Analysis"]
    header_row = next(
        r[0].row for r in ws.iter_rows(min_col=1, max_col=1) if r[0].value == "Fund"
    )
    assert ws.cell(row=header_row + 22, column=1).value is not None


def test_workbook_metadata_and_filename(report):
    wb = build_workbook(report)
    assert wb.properties.title == "Mutual Fund Portfolio Analysis"
    assert suggested_filename(date(2026, 9, 4)) == "portfolio-analysis-2026-09-04.xlsx"
